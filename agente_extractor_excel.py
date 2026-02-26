"""
AGENTE 2: Extraccion de Datos de Excel (Tareo Diario)
- Descarga adjuntos Excel de los correos de tareo
- Lee la pestaña "RESUMEN"
- Extrae: fecha del tareo, personal empresa, SINDICATO, TOTAL obreros, STAFF

Estructura real del Excel (pestaña RESUMEN):
  Fila 2: [B] titulo                          | [E] "Fecha:"
  Fila 3:                                     | [E] 2026-02-21 (datetime)
  Fila 5: [B] "DESCRIPCCION"                  | [E] "PERSONAL OBRERO" | [F] "PERSONAL STAFF"
  Fila 6: [B] "OPERARIO"                      | [E] 24               | [F] "---"
  Fila 7: [B] "OFICIAL"                       | [E] 15               | [F] "---"
  Fila 8: [B] "PEON"                          | [E] 19               | [F] "---"
  Fila 9: [B] "PERSONAL DE STAFF"             | [E] "---"            | [F] 19
  Fila 10:[B] "TOTAL"                         | [E] 58               | [F] 19

  Secciones de subcontratistas (variable):
  Fila N: [B] "CATEGORIA" | [C] "SUBCONTRATISTA1" | [D] "SUBCONTRATISTA2" | ...
  ...
  Fila M: [B] "TOTAL"     | [C] total_sub1        | [D] total_sub2        | ...

  Una de las secciones tiene la columna de la empresa (configurable)
  Una de las secciones puede tener "SINDICATO" como columna
"""
import io
import re
import base64
from datetime import date, datetime
from openpyxl import load_workbook

from config import EXCEL_COMPANY_KEYWORDS


def extraer_datos_tareo(gmail_service, mensaje_id, adjuntos, fecha_objetivo):
    """
    Extrae datos de personal del Excel adjunto al correo de tareo.

    Args:
        gmail_service: Gmail API service
        mensaje_id: ID del mensaje en Gmail
        adjuntos: Lista de adjuntos Excel [{filename, attachmentId}]
        fecha_objetivo: date object (fecha esperada del tareo)

    Returns:
        dict con datos extraidos
    """
    resultado = {
        "fecha_tareo": None,
        "fecha_correcta": False,
        "personal_empresa": 0,
        "personal_sindicato": 0,
        "total_obreros": 0,
        "personal_staff": 0,
        "fuente": None,
        "error": None,
        "formato_valido": False,
        "formato_detalles": [],
    }

    if not adjuntos:
        resultado["error"] = "Sin adjunto Excel"
        return resultado

    for adj in adjuntos:
        try:
            attachment = gmail_service.users().messages().attachments().get(
                userId="me", messageId=mensaje_id, id=adj["attachmentId"]
            ).execute()

            file_data = base64.urlsafe_b64decode(attachment["data"])
            datos = _procesar_excel_tareo(file_data, adj["filename"], fecha_objetivo)

            if datos:
                resultado.update(datos)
                resultado["fuente"] = adj["filename"]
                if datos.get("total_obreros", 0) > 0:
                    return resultado

        except Exception as e:
            resultado["error"] = f"Error con '{adj['filename']}': {e}"
            print(f"    [WARN] {resultado['error']}")

    if resultado["error"] is None:
        resultado["error"] = "No se pudo extraer datos del Excel"

    return resultado


def _procesar_excel_tareo(file_data, filename, fecha_objetivo):
    """Procesa un archivo Excel de tareo y extrae datos de la pestaña RESUMEN."""
    resultado = {
        "fecha_tareo": None,
        "fecha_correcta": False,
        "personal_empresa": 0,
        "personal_sindicato": 0,
        "total_obreros": 0,
        "personal_staff": 0,
        "formato_valido": False,
        "formato_detalles": [],
    }

    try:
        wb = load_workbook(io.BytesIO(file_data), data_only=True)
        sheet_names_lower = [s.lower().strip() for s in wb.sheetnames]

        # Buscar pestaña "RESUMEN"
        resumen_idx = None
        for i, name in enumerate(sheet_names_lower):
            if name == "resumen":
                resumen_idx = i
                break
        if resumen_idx is None:
            for i, name in enumerate(sheet_names_lower):
                if "resumen" in name:
                    resumen_idx = i
                    break

        if resumen_idx is None:
            print(f"    [WARN] No se encontro pestaña 'Resumen' en {filename}")
            resultado["formato_detalles"] = ["Pestaña RESUMEN no encontrada"]
            wb.close()
            return resultado

        ws = wb.worksheets[resumen_idx]
        max_row = min(ws.max_row or 100, 100)
        max_col = min(ws.max_column or 10, 10)

        # ==================================================================
        # 1. EXTRAER FECHA (buscar "Fecha:" y tomar el valor debajo o al lado)
        # ==================================================================
        for row_idx in range(1, min(10, max_row + 1)):
            for col_idx in range(1, max_col + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val and isinstance(cell_val, str) and "fecha" in cell_val.lower():
                    # Buscar datetime en celda de abajo (misma columna)
                    below = ws.cell(row=row_idx + 1, column=col_idx).value
                    if isinstance(below, (datetime, date)):
                        fecha = below.date() if isinstance(below, datetime) else below
                        resultado["fecha_tareo"] = fecha.strftime("%d/%m/%Y")
                        resultado["fecha_correcta"] = (fecha == fecha_objetivo)
                        break
                    # O al lado derecho
                    right = ws.cell(row=row_idx, column=col_idx + 1).value
                    if isinstance(right, (datetime, date)):
                        fecha = right.date() if isinstance(right, datetime) else right
                        resultado["fecha_tareo"] = fecha.strftime("%d/%m/%Y")
                        resultado["fecha_correcta"] = (fecha == fecha_objetivo)
                        break
            if resultado["fecha_tareo"]:
                break

        # ==================================================================
        # 2. EXTRAER TOTAL OBREROS Y STAFF (seccion superior fijas 5-15)
        #    Buscar fila con "TOTAL" en col B, leer col E y col F
        # ==================================================================
        for row_idx in range(5, min(15, max_row + 1)):
            cell_b = ws.cell(row=row_idx, column=2).value
            if cell_b and isinstance(cell_b, str) and cell_b.strip().upper() == "TOTAL":
                # Col E (5) = PERSONAL OBRERO total
                val_e = ws.cell(row=row_idx, column=5).value
                if isinstance(val_e, (int, float)) and val_e > 0:
                    resultado["total_obreros"] = int(val_e)
                # Col F (6) = PERSONAL STAFF total
                val_f = ws.cell(row=row_idx, column=6).value
                if isinstance(val_f, (int, float)) and val_f > 0:
                    resultado["personal_staff"] = int(val_f)
                break

        # ==================================================================
        # 3. EXTRAER personal empresa y SINDICATO de las secciones de subcontratistas
        #    Buscar filas con "CATEGORIA" en col B, luego leer headers de columnas
        #    y encontrar el TOTAL correspondiente
        # ==================================================================
        row_idx = 15  # Empezar despues de la seccion principal
        while row_idx <= max_row:
            cell_b = ws.cell(row=row_idx, column=2).value
            if cell_b and isinstance(cell_b, str) and cell_b.strip().upper() == "CATEGORIA":
                # Esta es una fila header de seccion de subcontratistas
                # Leer nombres de columnas (C, D, E, F, ...)
                col_map = {}
                for col_idx in range(3, max_col + 1):
                    header = ws.cell(row=row_idx, column=col_idx).value
                    if header and isinstance(header, str):
                        header_clean = header.strip().upper()
                        col_map[col_idx] = header_clean

                # Buscar la fila TOTAL de esta seccion (siguiente fila con "TOTAL" en col B)
                for total_row in range(row_idx + 1, min(row_idx + 15, max_row + 1)):
                    total_cell = ws.cell(row=total_row, column=2).value
                    if total_cell and isinstance(total_cell, str) and total_cell.strip().upper() == "TOTAL":
                        # Extraer valores para cada columna mapeada
                        for col_idx, header_name in col_map.items():
                            val = ws.cell(row=total_row, column=col_idx).value
                            total_val = int(val) if isinstance(val, (int, float)) and val > 0 else 0

                            if any(kw in header_name for kw in EXCEL_COMPANY_KEYWORDS):
                                # Sumar (puede haber variantes con "(SIND.)")
                                if "SIND" in header_name:
                                    resultado["personal_sindicato"] += total_val
                                else:
                                    resultado["personal_empresa"] += total_val

                            elif header_name == "SINDICATO":
                                resultado["personal_sindicato"] += total_val

                        break  # Encontramos TOTAL, pasar a siguiente seccion

            row_idx += 1

        # ==================================================================
        # 4. VALIDAR FORMATO OFICIAL (solo estructura de la pestaña RESUMEN)
        # ==================================================================
        fmt_valid, fmt_issues = _validar_formato_resumen(ws, max_row)
        resultado["formato_valido"] = fmt_valid
        resultado["formato_detalles"] = fmt_issues

        wb.close()

    except Exception as e:
        print(f"    [WARN] Error procesando Excel '{filename}': {e}")

    return resultado


def _validar_formato_resumen(ws, max_row):
    """
    Valida que la pestaña RESUMEN siga el formato oficial.
    Retorna (is_valid, issues).
    """
    issues = []

    # Check 1: Headers en fila 5
    cell_b5 = ws.cell(row=5, column=2).value
    cell_e5 = ws.cell(row=5, column=5).value
    has_headers = False
    if cell_b5 and isinstance(cell_b5, str) and "descrip" in cell_b5.lower():
        if cell_e5 and isinstance(cell_e5, str) and "obrero" in cell_e5.lower():
            has_headers = True
    if not has_headers:
        issues.append("Headers fila 5 no encontrados (DESCRIPCION, PERSONAL OBRERO)")

    # Check 2: Fila TOTAL con valores numericos entre filas 5-15
    has_total = False
    for r in range(5, min(15, max_row + 1)):
        cell_b = ws.cell(row=r, column=2).value
        if cell_b and isinstance(cell_b, str) and cell_b.strip().upper() == "TOTAL":
            val_e = ws.cell(row=r, column=5).value
            if isinstance(val_e, (int, float)):
                has_total = True
            break
    if not has_total:
        issues.append("Fila TOTAL no encontrada entre filas 5-15")

    # Check 3: Secciones CATEGORIA en filas 15+
    has_categoria = False
    for r in range(15, min(max_row + 1, 100)):
        cell_b = ws.cell(row=r, column=2).value
        if cell_b and isinstance(cell_b, str) and cell_b.strip().upper() == "CATEGORIA":
            has_categoria = True
            break
    if not has_categoria:
        issues.append("Secciones de subcontratistas (CATEGORIA) no encontradas")

    return len(issues) == 0, issues
